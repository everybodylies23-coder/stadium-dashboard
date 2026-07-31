import os
import sys
import json
import re
import datetime
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.cell import MergedCell
from bs4 import BeautifulSoup

# Force current working directory to be the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# Ensure input folder exists in the correct directory
os.makedirs("data_input", exist_ok=True)

CONFIG_FILE = "config.json"
RICH_SUMMARIES_FILE = "data_input/ai_summaries_rich.json"

def safe_save_workbook(wb, filepath):
    """
    Attempts to save the workbook. Bypasses openpyxl serialization bugs by converting 
    all ArrayFormula objects to normal raw formula strings (str) before saving.
    This prevents Excel from throwing sheet formula corruption warnings.
    """
    from openpyxl.worksheet.formula import ArrayFormula
    print("Optimizing formulas to prevent Excel corruption...")
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        converted = 0
        # Use iter_rows for extremely fast cell traversal without coordinate parsing overhead
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                val = cell.value
                if isinstance(val, ArrayFormula):
                    cell.value = val.text
                    converted += 1
        if converted > 0:
            print(f"  Converted {converted} ArrayFormulas in sheet '{sheetname}'")
            
    while True:
        try:
            wb.save(filepath)
            break
        except PermissionError:
            print(f"\n[エラー/警告] ファイルに書き込めません: '{filepath}'")
            print("Excelアプリ等でこのファイルが開いている可能性があります。")
            print("Excelを閉じてから、キーボードの【ENTER】キーを押して再試行してください。")
            input("閉じた後にENTERキーを押してください...")

def safe_write_cell(ws, row, col, value):
    """
    Safely writes a value to a cell, resolving the top-left cell if it is a MergedCell.
    """
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                top_left.value = value
                return top_left
    else:
        cell.value = value
        return cell

def is_empty_or_formula(val):
    """
    Determines if a cell value is empty, None, whitespace, or an Excel formula.
    """
    if val is None:
        return True
    val_str = str(val).strip()
    if not val_str:
        return True
    if val_str.startswith("="):
        return True
    return False

def normalize_date_string(date_val):
    """
    Normalizes any date object or string into standard 'YYYY/MM/DD' format.
    Handles 'YYYY-MM-DD HH:MM:SS', 'YYYY-MM-DD', etc.
    """
    if date_val is None:
        return ""
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime("%Y/%m/%d")
    if isinstance(date_val, datetime.date):
        return date_val.strftime("%Y/%m/%d")
        
    date_str = str(date_val).strip()
    match = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', date_str)
    if match:
        return f"{match.group(1)}/{int(match.group(2)):02d}/{int(match.group(3)):02d}"
    return date_str

def extract_excel_specific_days_setting(excel_path):
    """
    Extracts custom specific day end-digit settings from 【分析】高設定履歴DB Sheet Row 2.
    Example output: "末尾1, ゾロ目"
    """
    try:
        if not excel_path or not os.path.exists(excel_path):
            return "末尾1"
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "【分析】高設定履歴DB" in wb.sheetnames:
            ws = wb["【分析】高設定履歴DB"]
            for r in range(1, 10):
                c1_val = str(ws.cell(r, 1).value or '')
                if "特定日末尾" in c1_val or "特定日" in c1_val:
                    settings = []
                    for col in range(2, ws.max_column + 1):
                        val = ws.cell(r, col).value
                        if val is not None and str(val).strip() != '':
                            s_str = str(val).strip()
                            if s_str.isdigit():
                                settings.append(f"末尾{s_str}")
                            else:
                                settings.append(s_str)
                    wb.close()
                    if settings:
                        return ", ".join(settings)
        wb.close()
    except Exception as e:
        print(f"Notice: Could not read specific days setting from Excel: {e}")
    return "末尾1"


def extract_past_summaries(excel_path, days=30):
    """
    Excelの【AI】総括シートから、直近最大30日分のAI総括・店長心理ログを抽出する
    """
    try:
        if not excel_path or not os.path.exists(excel_path):
            return "（過去のAI総括ログなし）"
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        summaries = []
        if "【AI】総括" in wb.sheetnames:
            ws = wb["【AI】総括"]
            rows = []
            for r in range(2, ws.max_row + 1):
                dt_val = ws.cell(r, 5).value
                txt_val = ws.cell(r, 6).value
                if dt_val and txt_val:
                    dt_str = normalize_date_string(dt_val)
                    rows.append((dt_str, str(txt_val)))
            
            rows.sort(key=lambda x: x[0])
            recent_rows = rows[-days:]
            for dt_str, txt in recent_rows:
                summaries.append(f"【{dt_str} のAI総括・心理分析】\n{txt}\n")
        wb.close()
        return "\n".join(summaries) if summaries else "（過去のAI総括ログなし）"
    except Exception as e:
        print(f"Notice: Could not read past summaries: {e}")
        return "（過去のAI総括ログなし）"


def extract_top5_winning_areas(all_records):
    """
    全蓄積データから、設定スコア4.5以上（高設定挙動）の出現率が高い台番号TOP5を自動計算
    """
    from collections import defaultdict
    mach_scores = defaultdict(list)
    
    for r in all_records:
        m_num = r.get('machine_number')
        score = r.get('score', 0)
        if m_num:
            mach_scores[m_num].append(score)
            
    mach_win_rates = {}
    for m_num, scores in mach_scores.items():
        if len(scores) >= 2:
            high_count = sum(1 for s in scores if s >= 4.5)
            rate = high_count / len(scores)
            mach_win_rates[m_num] = (rate, len(scores), high_count)
            
    top_machines = sorted(mach_win_rates.items(), key=lambda x: (x[1][0], x[1][2]), reverse=True)[:5]
    
    lines = ["【🏆 店舗別・過去高設定勝率 台番号 TOP 5】"]
    if top_machines:
        for rank, (m_num, (rate, total, high)) in enumerate(top_machines, 1):
            lines.append(f"  第{rank}位: 台#{m_num} (高設定スコア率: {rate*100:.1f}% / 該当{high}回/{total}営業日)")
    else:
        lines.append("  （データ蓄積数不足のため現在算出中）")
        
    return "\n".join(lines)


def extract_manager_strategy_context(store_dir):
    """
    店舗フォルダ内に『過去の店長戦略コンテキスト.md』が存在する場合のみテキストを読み込み、
    存在しない場合は安全に無視する。
    """
    strategy_file = os.path.join(store_dir, "過去の店長戦略コンテキスト.md")
    if os.path.exists(strategy_file):
        try:
            with open(strategy_file, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if content:
                print("Notice: 『過去の店長戦略コンテキスト.md』を検出・読み込みました。")
                return f"■ セクション12: 過去の店長戦略コンテキスト（店舗固有の戦略・傾向知見メモ）\n{content}"
        except Exception as e:
            print(f"Notice: Could not read 過去の店長戦略コンテキスト.md: {e}")
            
    return "■ セクション12: 過去の店長戦略コンテキスト\n（※過去の店長戦略コンテキスト.md は配置されていないため無視します）"


def extract_event_vs_normal_merihari(all_records, excel_file):
    """
    全蓄積データと特定日設定を照合し、「特定日」と「通常日」での高設定（スコア4.5以上）投下台数のギャップを自動判定
    """
    try:
        from collections import defaultdict
        specific_days_rule = extract_excel_specific_days_setting(excel_file)
        
        rule_str = str(specific_days_rule).lower()
        
        records_by_date = defaultdict(list)
        for r in all_records:
            d_str = r.get('date')
            score = r.get('score', 0)
            if d_str:
                records_by_date[d_str].append(score)
                
        event_days_scores = []
        normal_days_scores = []
        
        for d_str, scores in records_by_date.items():
            try:
                dt_obj = datetime.datetime.strptime(d_str, "%Y/%m/%d")
                day_num = dt_obj.day
                month_num = dt_obj.month
            except ValueError:
                continue
                
            is_event = False
            digits = re.findall(r'\d+', rule_str)
            for d in digits:
                d_int = int(d)
                if d_int < 10 and day_num % 10 == d_int:
                    is_event = True
                elif day_num == d_int:
                    is_event = True
                    
            if "ゾロ目" in rule_str and (day_num in [11, 22] or day_num == month_num):
                is_event = True
                
            high_count = sum(1 for s in scores if s >= 4.5)
            if is_event:
                event_days_scores.append(high_count)
            else:
                normal_days_scores.append(high_count)
                
        avg_event_high = (sum(event_days_scores) / len(event_days_scores)) if event_days_scores else 0.0
        avg_normal_high = (sum(normal_days_scores) / len(normal_days_scores)) if normal_days_scores else 0.0
        
        lines = [
            "■ セクション13: 店舗の特定日 vs 通常日 メリハリ分析データ（自動判定）",
            f"  ・特定日設定: {specific_days_rule}",
            f"  ・特定日（平均高設定投下数）: {avg_event_high:.1f} 台 / 1営業日",
            f"  ・通常日（平均高設定投下数）: {avg_normal_high:.1f} 台 / 1営業日"
        ]
        
        if avg_event_high >= avg_normal_high * 2.0 and avg_normal_high <= 2.0:
            lines.append("  ⚠️ 【店舗診断: 強烈メリハリ型ホール】")
            lines.append("     この店舗は特定日に高設定が集中し、通常日は高設定が極めて少ない店舗です。")
            lines.append("     【最重要指示】: 通常日の分析では無理にS/Aランクの予想を出さず、評価を控えめ（Bランク中心または厳選1〜2台）にしてください。")
        else:
            lines.append("  ℹ️ 【店舗診断: 全日・ローテ型ホール】")
            lines.append("     この店舗は通常日にも一定数の高設定が投下されるホールです。ローテーションや履歴DBを重視して予想してください。")
            
        return "\n".join(lines)
    except Exception as e:
        print(f"Notice: Could not calculate event vs normal merihari: {e}")
        return "■ セクション13: 店舗の特定日 vs 通常日 メリハリ分析データ\n  （自動判定準備中）"


def load_config():
    if not os.path.exists(CONFIG_FILE):
        print(f"Error: {CONFIG_FILE} not found. Please create it first.")
        sys.exit(1)
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def read_file_with_encoding(filepath):
    encodings = ['utf-8', 'cp932', 'shift_jis', 'utf-8-sig']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
                if "機種" in content or "台番" in content or "差枚" in content:
                    print(f"File successfully read with encoding: {enc}")
                    return content
        except Exception:
            continue
            
    print("Warning: Could not detect standard encoding. Reading with fallback (errors ignored).")
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()

def parse_html_data(filepath):
    """
    Parses Pachislot data from a downloaded HTML file using header detection.
    Supports tables with or without '差枚' (difference coins) column.
    """
    results = []
    seen_machine_numbers = set()
    print(f"Parsing HTML data from: {filepath}")
    html_content = read_file_with_encoding(filepath)
    soup = BeautifulSoup(html_content, 'html.parser')
        
    tables = soup.find_all('table')
    for table in tables:
        headers = [th.get_text().strip() for th in table.find_all('th')]
        if not headers:
            first_row = table.find('tr')
            if first_row:
                headers = [td.get_text().strip() for td in first_row.find_all('td')]
                
        header_text = "".join(headers)
        
        is_slot_table = (
            ("BB" in header_text or "BIG" in header_text) and 
            ("RB" in header_text or "REG" in header_text) and 
            any(h in header_text for h in ["台番", "台番号"])
        )
        if not is_slot_table:
            continue
            
        has_diff = ("差枚" in header_text)
        
        rows = table.find_all('tr')
        for row in rows:
            cells = [td.get_text().strip() for td in row.find_all('td')]
            if len(cells) >= 4:
                cell_text = "".join(cells)
                if any(x in cell_text for x in ["平均", "累計", "合計", "台番号", "機種名"]):
                    continue
                try:
                    def clean_val(val):
                        return val.replace(",", "").replace("+", "").replace(" ", "").strip()
                        
                    c0_clean = clean_val(cells[0])
                    c1_clean = clean_val(cells[1])
                    
                    if c0_clean.isdigit():
                        machine_number = int(c0_clean)
                        machine_name = cells[1]
                        g_games = int(clean_val(cells[2]))
                        if has_diff:
                            diff_coins = int(clean_val(cells[3]))
                            bb_count = int(clean_val(cells[4]))
                            rb_count = int(clean_val(cells[5])) if len(cells) > 5 else 0
                        else:
                            diff_coins = 0
                            bb_count = int(clean_val(cells[3]))
                            rb_count = int(clean_val(cells[4])) if len(cells) > 4 else 0
                    else:
                        machine_name = cells[0]
                        machine_number = int(c1_clean)
                        g_games = int(clean_val(cells[2]))
                        if has_diff:
                            diff_coins = int(clean_val(cells[3]))
                            bb_count = int(clean_val(cells[4]))
                            rb_count = int(clean_val(cells[5])) if len(cells) > 5 else 0
                        else:
                            diff_coins = 0
                            bb_count = int(clean_val(cells[3]))
                            rb_count = int(clean_val(cells[4])) if len(cells) > 4 else 0
                    
                    if machine_number in seen_machine_numbers:
                        continue
                    seen_machine_numbers.add(machine_number)
                    
                    results.append({
                        "machine_name": machine_name,
                        "machine_number": machine_number,
                        "g_games": g_games,
                        "diff_coins": diff_coins,
                        "bb_count": bb_count,
                        "rb_count": rb_count
                    })
                except (ValueError, IndexError):
                    continue
                    
    print(f"Successfully extracted {len(results)} unique rows from HTML tables.")
    return results

def parse_text_data(filepath):
    results = []
    print(f"Parsing Text data from: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        tokens = re.split(r'\t+|\s{2,}', line)
        if len(tokens) < 4:
            tokens = line.split(' ')
            if len(tokens) < 4:
                continue

        if any(x in line for x in ["機種", "台番", "平均", "累計", "合計"]):
            continue

        try:
            numbers = []
            machine_name = ""
            for token in tokens:
                token = token.replace(",", "").replace("+", "").strip()
                if re.match(r'^-?\d+$', token):
                    numbers.append(int(token))
                else:
                    if token and token not in ["-", "★"]:
                        machine_name = token
            
            if len(numbers) >= 4:
                machine_number = numbers[0]
                g_games = numbers[1]
                diff_coins = numbers[2]
                bb_count = numbers[3]
                rb_count = len(numbers) > 4 and numbers[4] or 0
                
                results.append({
                    "machine_name": machine_name,
                    "machine_number": machine_number,
                    "g_games": g_games,
                    "diff_coins": diff_coins,
                    "bb_count": bb_count,
                    "rb_count": rb_count
                })
        except Exception as e:
            continue
            
    print(f"Successfully extracted {len(results)} rows from Text.")
    return results

def copy_excel_formula(ws, src_row, dst_row, max_col):
    """
    Copies formulas from src_row to dst_row for columns > 7, dynamically updating the row numbers.
    """
    from openpyxl.worksheet.formula import ArrayFormula
    for col in range(1, max_col + 1):
        cell = ws.cell(src_row, col)
        if col > 7:
            val = cell.value
            if isinstance(val, ArrayFormula):
                val = val.text
            if isinstance(val, str) and val.startswith("="):
                new_formula = re.sub(rf'\b([A-Z]+){src_row}\b', rf'\g<1>{dst_row}', val)
                ws.cell(dst_row, col, new_formula)
            elif val is not None:
                ws.cell(dst_row, col, val)

def update_excel_data(excel_path, target_date, parsed_data):
    import excel_calc_logic
    if not os.path.exists(excel_path):
        print(f"Error: Excel file {excel_path} not found.")
        sys.exit(1)
        
    print(f"Opening Excel file: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb['【データ】蓄積用']
    
    # Find actual max row ignoring formulas and empty rows
    real_max_row = 1
    for r in range(ws.max_row, 0, -1):
        val1 = ws.cell(r, 1).value
        val2 = ws.cell(r, 2).value
        if not is_empty_or_formula(val1) or not is_empty_or_formula(val2):
            real_max_row = r
            break
            
    max_col = ws.max_column
    
    # --- FORMULA AUTO-RESTORATION LOGIC ---
    # Python-based K-AF calculations completely bypass the need for Excel array formulas.
    # We no longer attempt to restore missing formulas in columns 11-32.
    # --------------------------------------
    
    # Check if target_date already exists in Column A (compare as dates)
    date_val = datetime.datetime.strptime(target_date, "%Y/%m/%d")
    date_exists = False
    for r in range(2, real_max_row + 1):
        cell_val = ws.cell(r, 1).value
        if isinstance(cell_val, datetime.datetime) and cell_val.date() == date_val.date():
            date_exists = True
            break
            
    if date_exists:
        print(f"Notice: Data for date {target_date} already exists in 【データ】蓄積用.")
        print("Skipping append to prevent duplication. Moving to AI analysis...")
        safe_save_workbook(wb, excel_path)  # Save the workbook safely
        return
        
    print(f"Current actual max row in 蓄積用: {real_max_row}. Appending {len(parsed_data)} lines...")
    
    # Pre-build specs, settings and history for Python K-AF calculation
    specs_dict, settings_dict = excel_calc_logic.build_maps(wb)
    processed_history = excel_calc_logic.extract_history(ws, real_max_row)
    
    start_row = real_max_row + 1
    for i, data in enumerate(parsed_data):
        current_row = start_row + i
        
        # Write values and force Date format on Column A
        dt_cell = ws.cell(current_row, 1, date_val)
        dt_cell.number_format = 'yyyy/mm/dd'
        
        ws.cell(current_row, 2, data["machine_name"])
        ws.cell(current_row, 3, data["machine_number"])
        ws.cell(current_row, 4, data["g_games"])
        ws.cell(current_row, 5, data["diff_coins"])
        ws.cell(current_row, 6, data["bb_count"])
        ws.cell(current_row, 7, data["rb_count"])
        
        # Calculate and write Big, Reg, and Combined probabilities (Cols 8, 9, 10)
        games_val = int(data["g_games"] or 0)
        bb_val = int(data["bb_count"] or 0)
        rb_val = int(data["rb_count"] or 0)
        ws.cell(current_row, 8, excel_calc_logic.format_probability(games_val, bb_val + rb_val))
        ws.cell(current_row, 9, excel_calc_logic.format_probability(games_val, bb_val))
        ws.cell(current_row, 10, excel_calc_logic.format_probability(games_val, rb_val))
        
        # Calculate K-AF columns via Python!
        row_dict = {
            "row_num": current_row,
            "date": date_val,
            "machine_name": str(data["machine_name"]) if data.get("machine_name") else "",
            "machine_number": int(data["machine_number"]),
            "g_games": int(data["g_games"] or 0),
            "diff_coins": int(data["diff_coins"] or 0),
            "bb_count": int(data["bb_count"] or 0),
            "rb_count": int(data["rb_count"] or 0)
        }
        excel_calc_logic.calculate_kaf_for_row(ws, current_row, row_dict, settings_dict, specs_dict, processed_history)

        
    print(f"Writing complete. Saving Excel file...")
    safe_save_workbook(wb, excel_path)
    print("Excel file successfully updated with new raw data and formulas.")

def prepare_ai_context(excel_path, target_date):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    data_ws = wb['【データ】蓄積用']
    
    all_records = []
    for r in range(2, data_ws.max_row + 1):
        d_val = data_ws.cell(r, 1).value
        m_name = data_ws.cell(r, 2).value
        m_num = data_ws.cell(r, 3).value
        g_val = data_ws.cell(r, 4).value
        diff_val = data_ws.cell(r, 5).value
        bb_val = data_ws.cell(r, 6).value
        rb_val = data_ws.cell(r, 7).value
        score_val = data_ws.cell(r, 29).value
        
        if d_val is None or m_num is None:
            continue
            
        if isinstance(d_val, (datetime.datetime, datetime.date)):
            d_str = d_val.strftime("%Y/%m/%d")
        else:
            d_str = str(d_val).strip().replace("-", "/")
            
        try:
            m_num_str = re.sub(r'\D', '', str(m_num).strip())
            m_num_int = int(m_num_str)
        except ValueError:
            continue
            
        g_int = int(g_val) if g_val is not None and str(g_val).isdigit() else 0
        diff_int = int(diff_val) if diff_val is not None and str(diff_val).replace("-","").isdigit() else 0
        try:
            sc_float = float(score_val) if score_val is not None else 0.0
        except ValueError:
            sc_float = 0.0
            
        all_records.append({
            'date': d_str,
            'machine_name': str(m_name or ""),
            'machine_number': m_num_int,
            'g_games': g_int,
            'diff_coins': diff_int,
            'bb': int(bb_val) if bb_val and str(bb_val).isdigit() else 0,
            'rb': int(rb_val) if rb_val and str(rb_val).isdigit() else 0,
            'score': sc_float
        })

    unique_dates = sorted(list(set(r['date'] for r in all_records)), reverse=True)
    recent_5_dates = unique_dates[-5:] if len(unique_dates) >= 5 else unique_dates

    # --- Section 1: End Digit Matrix ---
    tail_stats = {}
    for digit in range(10):
        tail_stats[digit] = {'total_diff': 0, 'total_count': 0, 'wins': 0, 'recent_diffs': []}
        
    for r in all_records:
        t_digit = r['machine_number'] % 10
        tail_stats[t_digit]['total_count'] += 1
        tail_stats[t_digit]['total_diff'] += r['diff_coins']
        if r['diff_coins'] > 0:
            tail_stats[t_digit]['wins'] += 1
            
    for d_str in recent_5_dates:
        day_recs = [r for r in all_records if r['date'] == d_str]
        for digit in range(10):
            d_recs = [r for r in day_recs if r['machine_number'] % 10 == digit]
            d_sum = sum(r['diff_coins'] for r in d_recs) if d_recs else 0
            tail_stats[digit]['recent_diffs'].append(f"{d_str[-5:]}:{d_sum:+d}枚")

    tail_matrix_lines = []
    for digit in range(10):
        st = tail_stats[digit]
        avg = round(st['total_diff'] / st['total_count']) if st['total_count'] > 0 else 0
        w_rate = round((st['wins'] / st['total_count']) * 100) if st['total_count'] > 0 else 0
        rec_str = " | ".join(st['recent_diffs'])
        tail_matrix_lines.append(f"末尾【{digit}】: 平均差枚 {avg:+d}枚 | 勝率 {w_rate}% ({st['wins']}/{st['total_count']}) | 直近推移: {rec_str}")

    # --- Section 2: Position Tendency Rates ---
    position_lines = []
    corner_recs = [r for r in all_records if r['machine_number'] % 10 in (1, 0, 5, 6)]
    corner_win = round((sum(1 for r in corner_recs if r['diff_coins'] > 0) / len(corner_recs)) * 100) if corner_recs else 0
    corner_avg = round(sum(r['diff_coins'] for r in corner_recs) / len(corner_recs)) if corner_recs else 0
    position_lines.append(f"- 角台/角付近台: 勝率 {corner_win}% | 平均差枚 {corner_avg:+d}枚")
    
    other_recs = [r for r in all_records if r['machine_number'] % 10 not in (1, 0, 5, 6)]
    other_win = round((sum(1 for r in other_recs if r['diff_coins'] > 0) / len(other_recs)) * 100) if other_recs else 0
    other_avg = round(sum(r['diff_coins'] for r in other_recs) / len(other_recs)) if other_recs else 0
    position_lines.append(f"- 中央/一般台: 勝率 {other_win}% | 平均差枚 {other_avg:+d}枚")

    # --- Section 3: Machine Brand Strength ---
    machine_summary = {}
    for r in all_records:
        name = r['machine_name']
        if name not in machine_summary:
            machine_summary[name] = {'total_diff': 0, 'count': 0, 'high_score_count': 0}
        machine_summary[name]['total_diff'] += r['diff_coins']
        machine_summary[name]['count'] += 1
        if r['score'] >= 4.5 or r['diff_coins'] >= 1500:
            machine_summary[name]['high_score_count'] += 1

    mach_lines = []
    for name, st in sorted(machine_summary.items(), key=lambda x: x[1]['total_diff'], reverse=True)[:15]:
        avg = round(st['total_diff'] / st['count']) if st['count'] > 0 else 0
        mach_lines.append(f"- {name}: 平均差枚 {avg:+d}枚 (全{st['count']}台) | 高設定/大爆発回数: {st['high_score_count']}回")

    # --- Section 4: Machine History Matrix ---
    mach_history_lines = []
    unique_mach_nums = sorted(list(set(r['machine_number'] for r in all_records)))
    for m_num in unique_mach_nums:
        m_recs = [r for r in all_records if r['machine_number'] == m_num]
        if not m_recs:
            continue
        m_recs = sorted(m_recs, key=lambda x: x['date'])[-5:]
        m_name = m_recs[-1]['machine_name']
        hist_str = ", ".join([f"{r['date'][-5:]}:G{r['g_games']}/差{r['diff_coins']:+d}/点{r['score']}" for r in m_recs])
        mach_history_lines.append(f"台#{m_num} ({m_name}): {hist_str}")

    # --- Section 5: High Setting History DB ---
    high_setting_db_lines = []
    for m_num in unique_mach_nums:
        m_recs = sorted([r for r in all_records if r['machine_number'] == m_num], key=lambda x: x['date'])
        high_recs = [r for r in m_recs if r['score'] >= 4.5 or r['diff_coins'] >= 1500]
        if high_recs:
            last_high_date = high_recs[-1]['date']
            events_since = len([r for r in m_recs if r['date'] > last_high_date])
            high_setting_db_lines.append(f"台#{m_num} ({m_recs[-1]['machine_name']}): 前回高設定 {last_high_date} | 経過イベント数 {events_since}回")

    # --- Section 6: AI Prediction Review History & Penalty Tracker ---
    pred_history_lines = []
    penalty_tracker = {}
    if "【AI】予想・答え合わせ" in wb.sheetnames:
        predict_ws = wb["【AI】予想・答え合わせ"]
        for r in range(4, predict_ws.max_row + 1):
            p_date = predict_ws.cell(r, 1).value
            p_name = predict_ws.cell(r, 2).value
            p_num = predict_ws.cell(r, 3).value
            p_reason = predict_ws.cell(r, 4).value
            p_g = predict_ws.cell(r, 5).value
            p_diff = predict_ws.cell(r, 6).value
            p_res = predict_ws.cell(r, 8).value
            
            if p_date is None or p_num is None:
                continue
                
            if isinstance(p_date, (datetime.datetime, datetime.date)):
                pd_str = p_date.strftime("%Y/%m/%d")
            else:
                pd_str = str(p_date).strip()
                
            try:
                m_num_str = re.sub(r'\D', '', str(p_num).strip())
                p_num_int = int(m_num_str)
            except ValueError:
                continue
                
            if p_num_int not in penalty_tracker:
                penalty_tracker[p_num_int] = {'total': 0, 'wins': 0, 'consecutive_losses': 0}
                
            penalty_tracker[p_num_int]['total'] += 1
            if str(p_res) == '〇':
                penalty_tracker[p_num_int]['wins'] += 1
                penalty_tracker[p_num_int]['consecutive_losses'] = 0
            elif str(p_res) == '×':
                penalty_tracker[p_num_int]['consecutive_losses'] += 1
                
            pred_history_lines.append(f"{pd_str} | 台#{p_num_int} ({p_name}) | 理由: {str(p_reason or '')[:25]} | G:{p_g} 差:{p_diff} | 判定:{p_res}")

    consecutive_loss_warnings = []
    for num, track in penalty_tracker.items():
        if track['consecutive_losses'] >= 2:
            consecutive_loss_warnings.append(f"⚠️ 台#{num}: 直近{track['consecutive_losses']}回連続で「×」判定中！(累計{track['wins']}/{track['total']}的中) ➔ 同一根拠での再推奨はペナルティ(-10〜-20点)適用！")

    # --- Section 7: User Confirmation Info ---
    confirm_data = []
    if "確認情報" in wb.sheetnames:
        confirm_ws = wb["確認情報"]
        for r in range(2, confirm_ws.max_row + 1):
            row_vals = [confirm_ws.cell(r, c).value for c in range(1, confirm_ws.max_column + 1)]
            if not any(v is not None for v in row_vals):
                continue
            r_str = " / ".join([str(v).strip() for v in row_vals if v is not None])
            if r_str:
                confirm_data.append(f"  - {r_str}")
    if not confirm_data:
        confirm_data.append("  (入力情報なし - データの周期性・ローテーション分析を主軸としてください)")

    # --- Section 8: Today's Highlighted Results (Juggler High Score & TOP 5 Performance) ---
    today_norm = str(target_date).strip().replace("-", "/")
    today_recs = [r for r in all_records if r['date'] == today_norm]
    if not today_recs and unique_dates:
        today_recs = [r for r in all_records if r['date'] == unique_dates[-1]]

    jug_keywords = ['ジャグラー', 'ジャグ', 'マイJ', 'ファンキー', 'アイム', 'ゴーゴー', 'ハッピー', 'ミラクル']
    juggler_high_scores = [r for r in today_recs if any(k in r['machine_name'] for k in jug_keywords) and r['score'] >= 4.5]
    juggler_high_score_lines = []
    for r in juggler_high_scores:
        juggler_high_score_lines.append(f"🌟 台#{r['machine_number']} ({r['machine_name']}): 最終スコア {r['score']} | G数 {r['g_games']}G | 差枚 {r['diff_coins']:+d}枚 | BB:{r['bb']} RB:{r['rb']}")

    top5_recs = sorted(today_recs, key=lambda x: x['diff_coins'], reverse=True)[:5]
    top5_areas_text = extract_top5_winning_areas(all_records)
    merihari_text = extract_event_vs_normal_merihari(all_records, excel_path)
    top5_lines = []
    for i, r in enumerate(top5_recs, 1):
        top5_lines.append(f"🏆 第{i}位 台#{r['machine_number']} ({r['machine_name']}): 差枚 {r['diff_coins']:+d}枚 | G数 {r['g_games']}G | スコア {r['score']}")

    wb.close()

    context_text = f"""
■ セクション1: 末尾別分析マトリクス（直近5開催）
{chr(10).join(tail_matrix_lines)}

■ セクション2: 台位置傾向割合
{chr(10).join(position_lines)}

■ セクション3: 機種別強さ・高設定投入回数
{chr(10).join(mach_lines)}

■ セクション4: 主要台番号別・直近5開催推移（G数/差枚/設定スコア）
{chr(10).join(mach_history_lines[:30])}

■ セクション5: 高設定履歴DB（前回高設定からの経過イベント数）
{chr(10).join(high_setting_db_lines[:25])}

■ セクション6: 過去AI予想答え合わせ＆連続×ペナルティ警告
【直近の答え合わせ履歴】
{chr(10).join(pred_history_lines[-20:])}

【連続×ペナルティ適用リスト（安易な再推奨厳禁）】
{chr(10).join(consecutive_loss_warnings) if consecutive_loss_warnings else "（現在、2回以上連続×の要警戒台はありません）"}

■ セクション7: ユーザー入力の店舗確認情報（公約・示唆・SNS等の一次情報）
{chr(10).join(confirm_data)}

■ セクション8: 本日（{target_date}）の注目実績（ダッシュボード連動表示）
【🌟 ジャグラー正解台一覧 (最終スコア4.5以上)】
{chr(10).join(juggler_high_score_lines) if juggler_high_score_lines else "（本日スコア4.5以上のジャグラー該当台なし）"}

【🏆 本日店舗実績 TOP 5 (差枚TOP5)】
{chr(10).join(top5_lines) if top5_lines else "（実績データなし）"}

■ セクション9: 店舗別・過去高設定勝率エリア TOP 5
{top5_areas_text}
"""
    return context_text


def run_gemini_analysis(api_key, context, target_date, excel_file=None):
    if excel_file and os.path.exists(excel_file):
        specific_days_rule = extract_excel_specific_days_setting(excel_file)
        past_summaries = extract_past_summaries(excel_file, days=30)
        store_dir = os.path.dirname(excel_file)
        manager_strategy_context = extract_manager_strategy_context(store_dir)
    else:
        specific_days_rule = "末尾5"
        past_summaries = "（過去のAI総括ログなし）"
        manager_strategy_context = """■ セクション12: 過去の店長戦略コンテキスト
（※過去の店長戦略コンテキスト.md は配置されていないため無視します）"""

    import google.generativeai as genai
    genai.configure(api_key=api_key)
    
    models_to_try = [
        'gemini-3.6-flash',
        'gemini-3.5-flash',
        'gemini-3.1-pro-preview',
        'gemini-2.5-pro',
        'gemini-flash-latest',
        'gemini-3.1-flash-lite',
        'gemini-2.0-flash',
        'gemini-2.0-flash-lite'
    ]
    
    d_obj = datetime.datetime.strptime(target_date, "%Y/%m/%d")
    tomorrow_date = (d_obj + datetime.timedelta(days=1)).strftime("%Y/%m/%d")
    
    prompt = f"""
あなたはパチスロホールの設定配分、店長心理、イベント周期、台番ローテーション、および行動経済学を解読するプロのデータサイエンティスト兼現役パチプロです。
提供された「Python自動集計による構造化統計データ」および「過去30日分のAI営業総括ログ」に基づき、分析対象日の【翌日（明日）：{tomorrow_date}】における推奨狙い台（計10台：ジャグラー系5台 ＋ スマスロ・その他5台）を定量スコアリングの上で決定し、営業総括を作成してください。

【分析対象店舗の特性】
店舗名：スタジアム店
- 分析対象店舗（スタジアム店）は毎日営業データをアナスロにて公開するデイルーホールです。
- 店舗の特定イベント日は【末尾5】です（Excel⚙️特定日末尾設定：{specific_days_rule}）。
- 分析対象日（本日：{target_date}）の【翌日（明日）：{tomorrow_date}】の高設定予想・狙い目台10台（ジャグラー5台＋スマスロ/その他5台）を決定してください。
- 本分テキスト内や予想根拠内に、未来の特定日（例：4月6日等）を狙い日付として絶対に記載せず、必ず【本日日付の翌日（明日）：{tomorrow_date}】を明記すること。

==================================================
【最重要：狙い目台選定の 7大細分化スコアリングルール（合計100点満点）】
==================================================
明日の推奨狙い台を選定する際、以下の7つの項目（合計100点満点）で厳密に加点・採点を行い、最終スコアを算出してください。

1. 【末尾・イベント周期との一致】（最大 20点）
   - 月日ゾロ目、日ゾロ目、特定末尾（例: 特定日末尾: {specific_days_rule}）との合致度。
   - セクション1の末尾マトリクスにおける高実績末尾へ強加点。

2. 【台番ローテーション・周期到達率】（最大 20点）
   - セクション5の高設定履歴DBにおける「前回高設定からの経過イベント数」およびローテーション周期到達度。

3. 【店長配分クセ・位置特性（勝率エリアTOP5）】（最大 15点）
   - セクション9の「過去高設定勝率エリア/台番 TOP 5」への該当度。
   - 角台、角2、島中央、隣スライド、据え置き傾向等の位置特性との一致。

4. 【直近不発からのお詫び・リベンジ投入期待】（最大 15点）
   - 直近で高稼働・高設定挙動を示しながら不発（差枚マイナス）に終わった台への「お詫びリセット」期待度。
   - セクション8の「本日スコア4.5以上台の投入パターン」との類似性。

5. 【機種扱いの強さ・本気度】（最大 15点）
   - セクション3の機種別高設定投入率、およびホールにおける看板機種（マイジャグ/主要スマスロ等）への優先度。
   - 過去30日間のAI総括から読み取れる店舗の主力推し機種傾向。

6. 【一次情報 ＆ 店舗固有戦略の一致】（最大 10点）
   - セクション7のユーザー入力（公約、確定系、LINE/SNS示唆等）および『セクション12: 過去の店長戦略コンテキスト』に記載された店長の長期戦略・配分癖メモと直接合致している場合、強力に補強・加点してください。

7. 【過去AI予想の答え合わせ精度・相性】（最大 5点）
   - セクション6の直近答え合わせ履歴。2回以上連続×の要警戒台（ペナルティ）を回避し、過去相性の良い台番を加点。

【ランク判定基準】
・推奨 S ランク: 90点 〜 100点（全条件が合致する超本命台）
・推奨 A ランク: 80点 〜 89点（強力な根拠が重複する対抗台）
・推奨 B ランク: 70点 〜 79点（データ・周期性に合致する抑え台）
※ 70点未満の台は推奨台として抽出しないこと。

==================================================
【最重要命令：ジャグラー高設定（スコア4.5以上）なぜなぜ分析 5 Whys】
==================================================
本日（{target_date}）の実績データにおいて、最終設定スコア 4.5 以上のジャグラー系該当台が存在する場合、営業総括テキストの冒頭に【🎯 ジャグラー高設定 なぜなぜ分析 5 Whys】セクションを設け、以下のフォーマットで5段階の因果関係を深掘りして言語化してください。

■ フォーマット例:
🎯 【ジャグラー高設定（スコア4.5以上）なぜなぜ分析 5 Whys】
・台#XXX（機種名 / スコア X.X）
  【Why 1: 直近要因】なぜこの台が高設定だったのか？（例: 前日大幅凹みからの上げ）
  【Why 2: 配分傾向】なぜそのパターンを選んだのか？（例: 3日連続不発台へのお詫びリセット）
  【Why 3: 時期要因】なぜ今日その施策を行ったのか？（例: 特定日前の事前稼働向上施策）
  【Why 4: 位置特性】なぜ他の候補ではなくこの場所か？（例: 中通路から見える角2視認エリア）
  【Why 5: 根本真因】なぜこの台番号だったのか？（例: 末尾Xへの事前示唆と店長勝負配分の合致）
  💡 【本日の配分法則の言語化・結論】店長の思考癖・配分パターンのまとめ

==================================================
【分析対象データ（コンテキスト）】
==================================================
{context}

■ セクション11: 過去30日分の【AI】営業総括・店長心理分析ログ
{past_summaries}

{manager_strategy_context}

==================================================
【出力フォーマット指示】
==================================================
以下の構成で出力してください。

### ① 【AI】営業評価と店長の心理総括
【次回（明日：{tomorrow_date}）の参戦評価：〇〇（例：狙い目だけ打ちに行く）】
次回予測日：{tomorrow_date}（イベント特徴：〇〇）

（※ここに「🎯 ジャグラー高設定 なぜなぜ分析 5 Whys」を挿入）

**店長心理分析・今後の配分傾向：**
（過去30日の総括ログおよび本日の結果を踏まえた深層心理の分析文章）

### ② 【AI】次回（明日：{tomorrow_date}）推奨狙い目台リスト（コピペ用）
| 日付 | 機種名 | 台番号 | 予想・狙い根拠 |
| :--- | :--- | :--- | :--- |
| {tomorrow_date} | マイジャグラーV | 590 | 【推奨Sランク - 95点】根拠：・・・ |
...（ジャグラー系5台 ＋ スマスロ・その他5台の計10台）
"""

    for model_name in models_to_try:
        try:
            print(f"Trying Gemini model: {model_name}...")
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            if response and response.text:
                print(f"Success with model: {model_name}!")
                return response.text
        except Exception as e:
            print(f"Model {model_name} failed: {e}")
            
    raise RuntimeError("All Gemini models failed to generate content.")


def write_ai_results_to_excel(excel_path, target_date, ai_text):
    d_obj = datetime.datetime.strptime(target_date, "%Y/%m/%d")
    
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # 1. Parse AI recommendations from Markdown table
    rows_to_insert = []
    lines = ai_text.splitlines()
    for line in lines:
        if "|" in line:
            cells = [c.strip() for c in line.split("|")]
            if len(cells) > 0 and cells[0] == "":
                cells.pop(0)
            if len(cells) > 0 and cells[-1] == "":
                cells.pop()
                
            if len(cells) >= 4:
                date_str = cells[0]
                machine_name = cells[1]
                try:
                    m_num_str = re.sub(r'\D', '', str(cells[2]).strip())
                    machine_num = int(m_num_str)
                except ValueError:
                    machine_num = cells[2]
                reason = cells[3]
                
                if "機種名" in machine_name or "---" in machine_name or "日付" in date_str:
                    continue
                    
                rows_to_insert.append((date_str, machine_name, machine_num, reason))
                
    if rows_to_insert:
        ai_ws = wb['【AI】予想・答え合わせ']
        
        first_date_str = rows_to_insert[0][0]
        try:
            pred_target_dt = datetime.datetime.strptime(first_date_str, "%Y/%m/%d")
        except ValueError:
            pred_target_dt = d_obj + datetime.timedelta(days=1)
            
        pred_target_str = pred_target_dt.strftime("%Y/%m/%d")
        
        # --- PREVENT DUPLICATE RECOMMENDATIONS ---
        rows_to_delete = []
        for r in range(4, ai_ws.max_row + 1):
            cell_val = ai_ws.cell(r, 1).value
            normalized_cell_val = normalize_date_string(cell_val)
            normalized_pred_target = normalize_date_string(pred_target_dt)
            if normalized_cell_val == normalized_pred_target:
                rows_to_delete.append(r)
                
        if rows_to_delete:
            print(f"Removing {len(rows_to_delete)} duplicate prediction rows in 【AI】予想・答え合わせ for date {pred_target_str}...")
            for r in reversed(rows_to_delete):
                ai_ws.delete_rows(r)
        # ------------------------------------------
        
        last_data_row = 3
        for r in range(ai_ws.max_row, 3, -1):
            val = ai_ws.cell(r, 1).value
            if not is_empty_or_formula(val):
                last_data_row = r
                break
        append_start_row = last_data_row + 1
            
        print(f"Writing {len(rows_to_insert)} recommendation rows for target date {pred_target_str} starting at Row {append_start_row}...")
        for i, row_data in enumerate(rows_to_insert):
            curr_r = append_start_row + i
            
            try:
                r_dt = datetime.datetime.strptime(row_data[0], "%Y/%m/%d")
            except ValueError:
                r_dt = pred_target_dt
                
            dt_cell = ai_ws.cell(curr_r, 1, r_dt)
            dt_cell.number_format = 'yyyy/mm/dd'
            
            ai_ws.cell(curr_r, 2, row_data[1])
            ai_ws.cell(curr_r, 3, row_data[2])
            ai_ws.cell(curr_r, 4, row_data[3])
    else:
        print("Warning: Could not parse recommendation table from AI response.")
        
    # 2. PERFORM PRE-CALCULATED LOOKUPS & AUTOMATIC UNMATCHED PREDICTION AUTO-LINKING
    print("Pre-calculating answer key data (lookup actual results)...")
    data_ws = wb['【データ】蓄積用']
    accumulated_db = {}
    data_dates = set()
    for r in range(2, data_ws.max_row + 1):
        date_val = data_ws.cell(r, 1).value
        mach_num = data_ws.cell(r, 3).value
        g_games = data_ws.cell(r, 4).value
        diff_coins = data_ws.cell(r, 5).value
        setting_score = data_ws.cell(r, 29).value  # Col 29 (AC): 最終設定スコア
        
        if date_val is not None:
            date_str = normalize_date_string(date_val)
            data_dates.add(date_str)
            try:
                m_num_str = re.sub(r'\D', '', str(mach_num).strip())
                m_num = int(m_num_str)
                accumulated_db[(date_str, m_num)] = (g_games, diff_coins, setting_score)
            except ValueError:
                continue
                
    ai_ws = wb['【AI】予想・答え合わせ']
    
    current_target_norm = normalize_date_string(d_obj)
    for r in range(4, ai_ws.max_row + 1):
        pred_date_val = ai_ws.cell(r, 1).value
        pred_g_val = ai_ws.cell(r, 5).value
        pred_diff_val = ai_ws.cell(r, 6).value
        
        if pred_date_val is not None and (pred_g_val is None or pred_g_val == "") and (pred_diff_val is None or pred_diff_val == ""):
            pred_date_norm = normalize_date_string(pred_date_val)
            if pred_date_norm not in data_dates and pred_date_norm <= current_target_norm:
                print(f"Auto-linking un-answered prediction Row {r} (was {pred_date_norm}) to current result date {current_target_norm}!")
                dt_c = ai_ws.cell(r, 1, d_obj)
                dt_c.number_format = 'yyyy/mm/dd'

    rewritten_count = 0
    for r in range(4, ai_ws.max_row + 1):
        date_val = ai_ws.cell(r, 1).value
        mach_val = ai_ws.cell(r, 3).value
        m_name = str(ai_ws.cell(r, 2).value or "")
        
        if date_val is not None and mach_val is not None:
            date_str = normalize_date_string(date_val)
            try:
                m_num_str = re.sub(r'\D', '', str(mach_val).strip())
                m_num = int(m_num_str)
                key = (date_str, m_num)
                if key in accumulated_db:
                    g_games, diff_coins, setting_score = accumulated_db[key]
                    
                    ai_ws.cell(r, 5, g_games)
                    ai_ws.cell(r, 6, diff_coins)
                    ai_ws.cell(r, 7, setting_score)
                    
                    is_jug = any(k in m_name for k in ['ジャグラー', 'ジャグ', 'マイJ', 'ファンキー', 'アイム', 'ゴーゴー', 'ハッピー', 'ミラクル'])
                    if is_jug:
                        if g_games is None or g_games == 0 or g_games == "":
                            res = "-"
                        else:
                            try:
                                sc_v = float(setting_score) if setting_score is not None else 0
                            except ValueError:
                                sc_v = 0
                            try:
                                df_v = int(diff_coins) if diff_coins is not None else -9999
                            except ValueError:
                                df_v = -9999
                            res = "〇" if (sc_v >= 4.5 and df_v >= 500) else "×"
                    else:
                        try:
                            df_v = int(diff_coins) if diff_coins is not None else -9999
                        except ValueError:
                            df_v = -9999
                        try:
                            gm_v = int(g_games) if g_games is not None else 0
                        except ValueError:
                            gm_v = 0
                        res = "〇" if (df_v >= 1000 or gm_v >= 6000) else "×"
                        
                    ai_ws.cell(r, 8, res)
                    rewritten_count += 1
            except ValueError:
                continue

    print(f"Pre-calculated & optimized {rewritten_count} cells in 【AI】予想・答え合わせ.")

    # 3. Create or Update Summary sheet (Consistently writing to Column E (5) & Column F (6))
    summary_ws = None
    if '【AI】総括' in wb.sheetnames:
        summary_ws = wb['【AI】総括']
    else:
        summary_ws = wb.create_sheet('【AI】総括')
        summary_ws.cell(1, 5, "日付 (結果日)")
        summary_ws.cell(1, 6, "AIからの総括 (コピペ用)")
        
    last_sum_r = 1
    for r in range(summary_ws.max_row, 0, -1):
        v5 = summary_ws.cell(r, 5).value
        v6 = summary_ws.cell(r, 6).value
        if not is_empty_or_formula(v5) or not is_empty_or_formula(v6):
            last_sum_r = r
            break
            
    found_date_row = None
    for r in range(1, last_sum_r + 1):
        c_val = summary_ws.cell(r, 5).value
        if normalize_date_string(c_val) == normalize_date_string(d_obj):
            found_date_row = r
            break
            
    target_sum_row = found_date_row if found_date_row else last_sum_r + 1
    if ai_text and ai_text.strip():
        dt_c = summary_ws.cell(target_sum_row, 5, d_obj)
        dt_c.number_format = 'yyyy/mm/dd'
        summary_ws.cell(target_sum_row, 6, ai_text)
        try:
            rich_summaries = {}
            if os.path.exists(RICH_SUMMARIES_FILE):
                with open(RICH_SUMMARIES_FILE, "r", encoding="utf-8") as f:
                    rich_summaries = json.load(f)
            rich_summaries[target_date] = ai_text
            with open(RICH_SUMMARIES_FILE, "w", encoding="utf-8") as f:
                json.dump(rich_summaries, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Warning: Could not update rich summaries file: {e}")
    
    safe_save_workbook(wb, excel_path)
    wb.close()
    print("AI Analysis results written successfully into sheets.")

def generate_html_dashboard(excel_path, store_name, has_diff_coins = True):
    """
    Reads data from the Excel workbook and generates a fully interactive, lightweight
    HTML dashboard with rich CSS styling, search filters, and statistics.
    This runs entirely in the browser with ZERO formula recalculation lag.
    """
    print(f"Generating HTML Interactive Dashboard for {store_name}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 1. Read Raw Slot Records (last 1500 records)
    records = []
    if "【データ】蓄積用" in wb.sheetnames:
        ws = wb["【データ】蓄積用"]
        real_max = 1
        for r in range(ws.max_row, 0, -1):
            val1 = ws.cell(r, 1).value
            val2 = ws.cell(r, 2).value
            if val1 is not None or val2 is not None:
                real_max = r
                break
        
        start_r = max(2, real_max - 15000)
        for r in range(start_r, real_max + 1):
            dt_val = ws.cell(r, 1).value
            dt_str = normalize_date_string(dt_val)
                
            records.append({
                "date": dt_str,
                "name": str(ws.cell(r, 2).value or ""),
                "number": int(ws.cell(r, 3).value or 0),
                "games": int(ws.cell(r, 4).value or 0),
                "diff": int(ws.cell(r, 5).value or 0),
                "bb": int(ws.cell(r, 6).value or 0),
                "rb": int(ws.cell(r, 7).value or 0),
                "comb_ratio": str(ws.cell(r, 8).value or "-"),
                "bb_ratio": str(ws.cell(r, 9).value or "-"),
                "rb_ratio": str(ws.cell(r, 10).value or "-"),
                "score": float(ws.cell(r, 29).value) if ws.cell(r, 29).value is not None and str(ws.cell(r, 29).value).replace('.','',1).isdigit() else 0.0
            })
            
    # 2. Read AI recommendations
    predictions = []
    if "【AI】予想・答え合わせ" in wb.sheetnames:
        ws = wb["【AI】予想・答え合わせ"]
        real_max = 3
        for r in range(ws.max_row, 3, -1):
            if ws.cell(r, 1).value is not None:
                real_max = r
                break
        for r in range(4, real_max + 1):
            dt_val = ws.cell(r, 1).value
            dt_str = normalize_date_string(dt_val)
                
            predictions.append({
                "date": dt_str,
                "name": str(ws.cell(r, 2).value or ""),
                "number": str(ws.cell(r, 3).value or ""),
                "reason": str(ws.cell(r, 4).value or ""),
                "games": ws.cell(r, 5).value,
                "diff": ws.cell(r, 6).value,
                "score": ws.cell(r, 7).value,
                "result": str(ws.cell(r, 8).value or "")
            })
            
    # 3. Read AI Summaries (Try rich paragraph file first, fallback to Excel)
    summaries = []
    rich_summaries = {}
    if os.path.exists(RICH_SUMMARIES_FILE):
        try:
            with open(RICH_SUMMARIES_FILE, "r", encoding="utf-8") as f:
                rich_summaries = json.load(f)
        except Exception:
            pass
            
    sum_sheet_name = None
    for name in wb.sheetnames:
        clean_name = name.replace(" ", "").replace("　", "")
        if "AI" in clean_name and "総括" in clean_name:
            sum_sheet_name = name
            break
            
    if sum_sheet_name:
        ws = wb[sum_sheet_name]
        real_max = 1
        for r in range(ws.max_row, 0, -1):
            if ws.cell(r, 5).value is not None or ws.cell(r, 6).value is not None:
                real_max = r
                break
        for r in range(1, real_max + 1):
            dt_val = ws.cell(r, 5).value
            dt_str = normalize_date_string(dt_val)
                
            val_f = ws.cell(r, 6).value
            if dt_str or val_f:
                summary_text = rich_summaries.get(dt_str, str(val_f or ""))
                summaries.append({
                    "date": dt_str,
                    "text": summary_text
                })
                
    wb.close()
    
    # Generate HTML content with updated 1-column layout, marked.js, and segmented tables
    html_template = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>【{store_name}】データ分析 AIダッシュボード</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.0.0"></script>
    <script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&family=Noto+Sans+JP:wght@300;400;700&display=swap" rel="stylesheet">
    <style>
        body {{
            font-family: 'Outfit', 'Noto Sans JP', sans-serif;
            background-color: #0f172a;
            color: #f8fafc;
        }}
        .glass {{
            background: rgba(30, 41, 59, 0.7);
            backdrop-filter: blur(12px);
            border: 1px solid rgba(255, 255, 255, 0.05);
        }}
        .prose h1, .prose h2, .prose h3 {{
            color: #ffffff;
            font-weight: 700;
            margin-top: 1.25em;
            margin-bottom: 0.5em;
        }}
        .prose h2 {{ font-size: 1.25rem; border-left: 3px solid #10b981; padding-left: 0.5rem; }}
        .prose p {{ margin-bottom: 1em; color: #cbd5e1; font-size: 0.875rem; }}
        .prose ul {{ list-style-type: disc; padding-left: 1.25rem; margin-bottom: 1em; color: #cbd5e1; font-size: 0.875rem; }}
        .prose li {{ margin-bottom: 0.25rem; }}
        .prose strong {{ color: #f43f5e; font-weight: 600; }}
    </style>
</head>
<body class="min-h-screen p-4 md:p-8">
    <div class="max-w-4xl mx-auto space-y-6">
        
        <!-- Header -->
        <header class="flex flex-col md:flex-row justify-between items-start md:items-center p-6 glass rounded-2xl gap-4">
            <div>
                <h1 class="text-3xl font-bold bg-gradient-to-r from-emerald-400 to-cyan-400 bg-clip-text text-transparent">
                    {store_name} 分析ダッシュボード
                </h1>
                <p class="text-slate-400 text-sm mt-1">AI Data Science & Prediction Suite</p>
            </div>
            <div class="flex items-center gap-3">
                <label for="date-select" class="text-sm text-slate-400">表示日付:</label>
                <select id="date-select" class="bg-slate-800 border border-slate-700 text-white rounded-lg px-3 py-2 focus:ring-2 focus:ring-emerald-400 focus:outline-none">
                    <!-- Dates populated dynamically -->
                </select>
            </div>
        </header>

        <!-- Stats Overview Grid -->
        <div class="grid grid-cols-2 md:grid-cols-4 gap-4">
            <div class="p-4 md:p-6 glass rounded-2xl text-center md:text-left">
                <h3 class="text-slate-400 text-xs font-semibold uppercase tracking-wider" id="stat-card1-title">本日総差枚</h3>
                <p id="stat-total-diff" class="text-xl md:text-2xl font-bold mt-2 text-white">0 枚</p>
            </div>
            <div class="p-4 md:p-6 glass rounded-2xl text-center md:text-left">
                <h3 class="text-slate-400 text-xs font-semibold uppercase tracking-wider" id="stat-card2-title">本日平均差枚</h3>
                <p id="stat-avg-diff" class="text-xl md:text-2xl font-bold mt-2 text-white">0 枚</p>
            </div>
            <div class="p-4 md:p-6 glass rounded-2xl text-center md:text-left">
                <h3 class="text-slate-400 text-xs font-semibold uppercase tracking-wider">本日勝率</h3>
                <p id="stat-win-rate" class="text-xl md:text-2xl font-bold mt-2 text-white">0%</p>
            </div>
            <div class="p-4 md:p-6 glass rounded-2xl text-center md:text-left">
                <h3 class="text-slate-400 text-xs font-semibold uppercase tracking-wider">AI予測 累積勝率(対象日迄)</h3>
                <p id="stat-ai-accuracy" class="text-xl md:text-2xl font-bold mt-2 text-emerald-400">0%</p>
            </div>
        </div>

        <!-- ① AIによる営業総括・心理分析 -->
        <section class="p-6 glass rounded-2xl space-y-4">
            <div class="flex items-center gap-2">
                <span class="p-1.5 bg-emerald-500/10 text-emerald-400 rounded-lg text-sm">💡</span>
                <h2 class="text-xl font-bold">AIによる営業総括・心理分析</h2>
            </div>
            <!-- Markdown parsed and loaded dynamically -->
            <div id="ai-summary-text" class="prose max-w-none p-4 bg-slate-900/50 rounded-xl border border-slate-800">
                該当日の総括データがありません。
            </div>
        </section>

        <!-- ② 主要機種の差枚状況ランキング (グラフ) -->
        <section class="p-6 glass rounded-2xl space-y-4">
            <div class="flex items-center gap-2">
                <span class="p-1.5 bg-indigo-500/10 text-indigo-400 rounded-lg text-sm">📊</span>
                <h2 class="text-xl font-bold font-semibold" id="chart-section-title">主要機種の差枚状況ランキング（平均値）</h2>
            </div>
            <div class="h-80 md:h-96 w-full">
                <canvas id="chart-machines"></canvas>
            </div>
        </section>

        <!-- ③ 次回（明日）のAI推奨台 (Segmented) -->
        <section class="p-6 glass rounded-2xl space-y-4">
            <div class="flex items-center gap-2">
                <span class="p-1.5 bg-cyan-500/10 text-cyan-400 rounded-lg text-sm">🎯</span>
                <h2 class="text-xl font-bold">次回（明日）のAI推奨台</h2>
            </div>
            
            <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
                <!-- Juggler Section -->
                <div class="space-y-2">
                    <h3 class="text-emerald-400 font-bold text-sm border-b border-emerald-500/30 pb-1">🤡 ジャグラー系 (5台)</h3>
                    <div class="overflow-x-auto">
                        <table class="w-full text-left border-collapse text-xs">
                            <thead>
                                <tr class="border-b border-slate-800 text-slate-400">
                                    <th class="py-2 px-2">機種名</th>
                                    <th class="py-2 px-2">台番</th>
                                    <th class="py-2 px-2 text-center">ランク</th>
                                    <th class="py-2 px-2">予想・狙い根拠</th>
                                </tr>
                            </thead>
                            <tbody id="pred-juggler-body" class="divide-y divide-slate-800/50">
                                <!-- Populated dynamically -->
                            </tbody>
                        </table>
                    </div>
                </div>

                <!-- Others/Smaslot Section -->
                <div class="space-y-2">
                    <h3 class="text-cyan-400 font-bold text-sm border-b border-cyan-500/30 pb-1">⚡ スマスロ・その他 (5台)</h3>
                    <div class="overflow-x-auto">
                        <table class="w-full text-left border-collapse text-xs">
                            <thead>
                                <tr class="border-b border-slate-800 text-slate-400">
                                    <th class="py-2 px-2">機種名</th>
                                    <th class="py-2 px-2">台番</th>
                                    <th class="py-2 px-2 text-center">ランク</th>
                                    <th class="py-2 px-2">予想・狙い根拠</th>
                                </tr>
                            </thead>
                            <tbody id="pred-others-body" class="divide-y divide-slate-800/50">
                                <!-- Populated dynamically -->
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </section>

        <!-- ④ AI予測答え合わせ ＆ 当日実績分析 -->
        <section class="p-6 glass rounded-2xl space-y-6">
            <div class="flex items-center gap-2">
                <span class="p-1.5 bg-amber-500/10 text-amber-400 rounded-lg text-sm">📈</span>
                <h2 class="text-xl font-bold">AI予測答え合わせ ＆ 店舗実績検証</h2>
            </div>
            
            <!-- ROW 1: Juggler Answer (Left) vs Juggler High Setting List Score >= 4.5 (Right) -->
            <div class="space-y-2">
                <h3 class="text-emerald-400 font-bold text-base border-b border-emerald-500/30 pb-1 flex items-center justify-between">
                    <span>🤡 ジャグラー系 検証</span>
                    <span class="text-xs font-normal text-slate-400">（左: AI予想結果 / 右: 実際の高設定挙動台）</span>
                </h3>
                <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
                    <!-- Juggler AI Prediction Answers -->
                    <div class="bg-slate-900/40 p-4 rounded-xl border border-slate-800 space-y-2">
                        <h4 class="text-xs font-semibold text-emerald-300">【AI予測 答え合わせ】</h4>
                        <div class="overflow-x-auto">
                            <table class="w-full text-left border-collapse text-xs">
                                <thead>
                                    <tr class="border-b border-slate-800 text-slate-400">
                                        <th class="py-2 px-1">台番 (機種)</th>
                                        <th class="py-2 px-1 text-center">ランク</th>
                                        <th class="py-2 px-1 text-right">${'差枚' if has_diff_coins else 'G数'}</th>
                                        <th class="py-2 px-1 text-center">スコア</th>
                                        <th class="py-2 px-1 text-center">結果</th>
                                    </tr>
                                </thead>
                                <tbody id="ans-juggler-body" class="divide-y divide-slate-800/50">
                                    <!-- Populated dynamically -->
                                </tbody>
                            </table>
                        </div>
                    </div>

                    <!-- Juggler Actual High Setting List (Score >= 4.5) -->
                    <div class="bg-slate-900/40 p-4 rounded-xl border border-slate-800 space-y-2">
                        <h4 class="text-xs font-semibold text-amber-300 flex items-center justify-between">
                            <span>【ジャグラー正解台一覧】(最終スコア 4.5以上)</span>
                            <span id="jug-high-count" class="text-[10px] bg-amber-500/20 text-amber-300 px-2 py-0.5 rounded-full">0台</span>
                        </h4>
                        <div class="overflow-x-auto">
                            <table class="w-full text-left border-collapse text-xs">
                                <thead>
                                    <tr class="border-b border-slate-800 text-slate-400">
                                        <th class="py-2 px-1">台番 (機種)</th>
                                        <th class="py-2 px-1 text-right">G数 (REG)</th>
                                        <th class="py-2 px-1 text-right">${'差枚' if has_diff_coins else '合算'}</th>
                                        <th class="py-2 px-1 text-center">スコア</th>
                                    </tr>
                                </thead>
                                <tbody id="juggler-high-score-body" class="divide-y divide-slate-800/50">
                                    <!-- Populated dynamically -->
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <!-- ROW 2: Other/Smart Slot Answer (Left) vs Store Top 5 Performance (Right) -->
            <div class="space-y-2 pt-2 border-t border-slate-800/60">
                <h3 class="text-cyan-400 font-bold text-base border-b border-cyan-500/30 pb-1 flex items-center justify-between">
                    <span>⚡ スマスロ・その他 検証 ＆ TOP5</span>
                    <span class="text-xs font-normal text-slate-400">（左: AI予想結果 / 右: 当日実績TOP5）</span>
                </h3>
                <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
                    <!-- Others AI Prediction Answers -->
                    <div class="bg-slate-900/40 p-4 rounded-xl border border-slate-800 space-y-2">
                        <h4 class="text-xs font-semibold text-cyan-300">【AI予測 答え合わせ】</h4>
                        <div class="overflow-x-auto">
                            <table class="w-full text-left border-collapse text-xs">
                                <thead>
                                    <tr class="border-b border-slate-800 text-slate-400">
                                        <th class="py-2 px-1">台番 (機種)</th>
                                        <th class="py-2 px-1 text-center">ランク</th>
                                        <th class="py-2 px-1 text-right">${'差枚' if has_diff_coins else 'G数'}</th>
                                        <th class="py-2 px-1 text-center">スコア</th>
                                        <th class="py-2 px-1 text-center">結果</th>
                                    </tr>
                                </thead>
                                <tbody id="ans-others-body" class="divide-y divide-slate-800/50">
                                    <!-- Populated dynamically -->
                                </tbody>
                            </table>
                        </div>
                    </div>

                    <!-- Store Top 5 Performance -->
                    <div class="bg-slate-900/40 p-4 rounded-xl border border-slate-800 space-y-2">
                        <h4 class="text-xs font-semibold text-indigo-300" id="top-5-title">
                            【${'本日差枚 TOP5' if has_diff_coins else '本日回転数 TOP5'}】
                        </h4>
                        <div class="overflow-x-auto">
                            <table class="w-full text-left border-collapse text-xs">
                                <thead>
                                    <tr class="border-b border-slate-800 text-slate-400">
                                        <th class="py-2 px-1">順位 / 台番 (機種)</th>
                                        <th class="py-2 px-1 text-right">G数</th>
                                        <th class="py-2 px-1 text-right">${'差枚' if has_diff_coins else 'スコア'}</th>
                                    </tr>
                                </thead>
                                <tbody id="top-performance-body" class="divide-y divide-slate-800/50">
                                    <!-- Populated dynamically -->
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </section>

        <!-- ⑤ 本日（選択日）の全台データ一覧 -->
        <section class="p-6 glass rounded-2xl space-y-4">
            <div class="flex flex-col md:flex-row justify-between items-start md:items-center gap-4">
                <div class="flex items-center gap-2">
                    <span class="p-1.5 bg-slate-500/10 text-slate-400 rounded-lg text-sm">📋</span>
                    <h2 class="text-xl font-bold">本日（選択日）の全台データ一覧</h2>
                </div>
                <input type="text" id="search-input" placeholder="機種名や台番号で検索..." class="w-full md:w-64 bg-slate-800 border border-slate-700 text-white rounded-lg px-3 py-1.5 focus:ring-2 focus:ring-emerald-400 focus:outline-none text-sm">
            </div>
            <div class="overflow-x-auto max-h-96 overflow-y-auto">
                <table class="w-full text-left border-collapse text-sm">
                    <thead class="sticky top-0 bg-slate-900 border-b border-slate-800 text-slate-400 z-10">
                        <tr>
                            <th class="py-3 px-2">台番号</th>
                            <th class="py-3 px-2">機種名</th>
                            <th class="py-3 px-2">G数</th>
                            <th class="py-3 px-2">差枚</th>
                            <th class="py-3 px-2">BB</th>
                            <th class="py-3 px-2">RB</th>
                            <th class="py-3 px-2">合算</th>
                            <th class="py-3 px-2">BB確率</th>
                            <th class="py-3 px-2">REG確率</th>
                        </tr>
                    </thead>
                    <tbody id="raw-data-table-body" class="divide-y divide-slate-800/50">
                        <!-- Populated dynamically -->
                    </tbody>
                </table>
            </div>
        </section>

    </div>

    <!-- Embedding JSON Data -->
    <script>
        const HAS_DIFF_COINS = {'true' if has_diff_coins else 'false'};
        const rawRecords = {json.dumps(records, ensure_ascii=False)};
        const rawPredictions = {json.dumps(predictions, ensure_ascii=False)};
        const rawSummaries = {json.dumps(summaries, ensure_ascii=False)};
        
        let machineChart = null;
        if (typeof Chart !== 'undefined' && typeof ChartDataLabels !== 'undefined') {{
            try {{
                Chart.register(ChartDataLabels);
            }} catch (e) {{
                console.warn('ChartDataLabels warning:', e);
            }}
        }}

        document.addEventListener('DOMContentLoaded', () => {{
            initializeDashboard();
        }});

        function initializeDashboard() {{
            const dates = [...new Set(rawRecords.map(r => r.date))].sort().reverse();
            const dateSelect = document.getElementById('date-select');
            
            dates.forEach(d => {{
                const opt = document.createElement('option');
                opt.value = d;
                opt.textContent = d;
                dateSelect.appendChild(opt);
            }});

            dateSelect.addEventListener('change', (e) => {{
                updateDashboardForDate(e.target.value);
            }});

            if (dates.length > 0) {{
                const latestDate = dates[dates.length - 1];
                dateSelect.value = latestDate;
                updateDashboardForDate(latestDate);
            }}
            
            calculateOverallAIAccuracy(targetDate);
        }}

                                        function extractRankBadge(reasonStr) {{
            if (!reasonStr) return '<span class="px-2 py-0.5 bg-slate-700 text-slate-300 rounded font-bold text-[11px]">-</span>';
            const match = reasonStr.match(/(?:推奨)?([SAB])(?:ランク)?/i);
            if (match) {{
                const rank = match[1].toUpperCase();
                if (rank === 'S') return '<span class="px-2 py-0.5 bg-rose-500/20 text-rose-400 border border-rose-500/30 rounded-full font-bold text-[11px]">S</span>';
                if (rank === 'A') return '<span class="px-2 py-0.5 bg-amber-500/20 text-amber-400 border border-amber-500/30 rounded-full font-bold text-[11px]">A</span>';
                if (rank === 'B') return '<span class="px-2 py-0.5 bg-blue-500/20 text-blue-400 border border-blue-500/30 rounded-full font-bold text-[11px]">B</span>';
            }}
            return '<span class="px-2 py-0.5 bg-slate-700 text-slate-300 rounded font-bold text-[11px]">-</span>';
        }}

        function isJugglerMachine(name) {{
            const jugKeywords = ['ジャグラー', 'ジャグ', 'マイJ', 'ファンキー', 'アイム', 'ゴーゴー', 'ハッピー', 'ミラクル'];
            return jugKeywords.some(keyword => name.includes(keyword));
        }}

        // Deduplicates records by taking the LAST registered prediction for a unique (date, number)
        function deduplicatePredictions(preds) {{
            const uniqueMap = {{}};
            preds.forEach(p => {{
                const key = p.date + '_' + p.number;
                uniqueMap[key] = p; // Keep the latest one
            }});
            return Object.values(uniqueMap);
        }}

        function updateDashboardForDate(targetDate) {{
            const filteredRecords = rawRecords.filter(r => r.date === targetDate);
            
            // 1. Update Stat boxes
            const totalDiff = filteredRecords.reduce((acc, curr) => acc + curr.diff, 0);
            const avgDiff = filteredRecords.length > 0 ? Math.round(totalDiff / filteredRecords.length) : 0;
            const winningSlots = filteredRecords.filter(r => r.diff > 0).length;
            const winRate = filteredRecords.length > 0 ? Math.round((winningSlots / filteredRecords.length) * 100) : 0;

            document.getElementById('stat-total-diff').textContent = totalDiff.toLocaleString() + ' 枚';
            document.getElementById('stat-total-diff').className = 'text-xl md:text-2xl font-bold mt-2 ' + (totalDiff >= 0 ? 'text-emerald-400' : 'text-rose-400');
            document.getElementById('stat-avg-diff').textContent = avgDiff.toLocaleString() + ' 枚';
            document.getElementById('stat-win-rate').textContent = winRate + '%';
            calculateOverallAIAccuracy(targetDate);

            // 2. Load AI Summary text (Markdown render full text, filtering out raw table lines)
            const summaryObj = rawSummaries.find(s => s.date === targetDate);
            if (summaryObj && summaryObj.text) {{
                let cleanText = summaryObj.text;
                const lines = cleanText.split('\\n');
                const filteredLines = lines.filter(line => {{
                    const trimmed = line.trim();
                    if (trimmed.startsWith('|') && trimmed.endsWith('|')) return false;
                    if (trimmed.includes('⑤ 【AI】予想・答え合わせ') || trimmed.includes('コピペ用テーブル')) return false;
                    return true;
                }});
                cleanText = filteredLines.join('\\n').trim();
                
                document.getElementById('ai-summary-text').innerHTML = marked.parse(cleanText);
            }} else {{
                document.getElementById('ai-summary-text').innerHTML = '<p class="text-slate-500">この日のAI営業総括データはありません。</p>';
            }}
            // 3. Load Predictions for the next event day (Segmented Juggler vs Others & Deduplicated!)
            const dedupedPredictions = deduplicatePredictions(rawPredictions);
            const predictionDates = [...new Set(dedupedPredictions.map(p => p.date))].sort();
            let nextEventDate = predictionDates.find(dStr => dStr > targetDate);
            if (!nextEventDate && predictionDates.length > 0) {{
                nextEventDate = predictionDates[predictionDates.length - 1];
            }}
            
            const nextEventTitle = document.getElementById('next-event-title');
            if (nextEventTitle && nextEventDate) {{
                nextEventTitle.textContent = `次回（${{nextEventDate}}）のAI推奨狙い目台`;
            }} else if (nextEventTitle) {{
                nextEventTitle.textContent = '次回イベント推奨台';
            }}
            
            const filteredPredictions = nextEventDate ? dedupedPredictions.filter(p => p.date === nextEventDate) : [];
            
            const predJugglerBody = document.getElementById('pred-juggler-body');
            const predOthersBody = document.getElementById('pred-others-body');
            predJugglerBody.innerHTML = '';
            predOthersBody.innerHTML = '';
            
            const jugPredictions = filteredPredictions.filter(p => isJugglerMachine(p.name)).slice(0, 5);
            const otherPredictions = filteredPredictions.filter(p => !isJugglerMachine(p.name)).slice(0, 5);
            
            if (jugPredictions.length === 0) {{
                predJugglerBody.innerHTML = '<tr><td colspan="4" class="py-3 text-center text-slate-500">ジャグラー推奨台はありません。</td></tr>';
            }} else {{
                jugPredictions.forEach(p => {{
                    const tr = document.createElement('tr');
                    tr.className = 'hover:bg-slate-800/30';
                    const rankBadge = extractRankBadge(p.reason);
                    tr.innerHTML = `
                        <td class="py-2.5 px-2 font-semibold text-slate-200">${{p.name}}</td>
                        <td class="py-2.5 px-2 text-emerald-400 font-bold font-mono">${{p.number}}</td>
                        <td class="py-2.5 px-2 text-center">${{rankBadge}}</td>
                        <td class="py-2.5 px-2 text-xs text-slate-300">${{p.reason}}</td>
                    `;
                    predJugglerBody.appendChild(tr);
                }});
            }}

            if (otherPredictions.length === 0) {{
                predOthersBody.innerHTML = '<tr><td colspan="4" class="py-3 text-center text-slate-500">スマスロ・その他推奨台はありません。</td></tr>';
            }} else {{
                otherPredictions.forEach(p => {{
                    const tr = document.createElement('tr');
                    tr.className = 'hover:bg-slate-800/30';
                    const rankBadge = extractRankBadge(p.reason);
                    tr.innerHTML = `
                        <td class="py-2.5 px-2 font-semibold text-slate-200">${{p.name}}</td>
                        <td class="py-2.5 px-2 text-cyan-400 font-bold font-mono">${{p.number}}</td>
                        <td class="py-2.5 px-2 text-center">${{rankBadge}}</td>
                        <td class="py-2.5 px-2 text-xs text-slate-300">${{p.reason}}</td>
                    `;
                    predOthersBody.appendChild(tr);
                }});
            }}

            // 4. Load Answers for targetDate predictions (Segmented & Deduplicated!)
            const filteredAnswers = dedupedPredictions.filter(p => p.date === targetDate);
            
            const ansJugglerBody = document.getElementById('ans-juggler-body');
            const ansOthersBody = document.getElementById('ans-others-body');
            ansJugglerBody.innerHTML = '';
            ansOthersBody.innerHTML = '';
            
            const jugAnswers = filteredAnswers.filter(a => isJugglerMachine(a.name)).slice(0, 5);
            const otherAnswers = filteredAnswers.filter(a => !isJugglerMachine(a.name)).slice(0, 5);
            
            const renderAnswerRow = (a, isJuggler) => {{
                const diffVal = a.diff !== null && a.diff !== undefined ? parseInt(a.diff) : null;
                const diffText = diffVal !== null ? (diffVal >= 0 ? '+' + diffVal : diffVal) + ' 枚' : '-';
                const diffClass = diffVal !== null ? (diffVal >= 0 ? 'text-emerald-400 font-bold' : 'text-rose-400') : 'text-slate-400';
                
                let scoreText = '-';
                if (a.score !== null && a.score !== undefined && a.score !== "") {{
                    const val = parseFloat(a.score);
                    scoreText = isNaN(val) ? String(a.score) : val.toFixed(1);
                }}
                
                let finalResult = a.result;
                if (isJuggler) {{
                    if (a.games === null || a.games === undefined || a.games === "" || a.games === 0) {{
                        finalResult = '-';
                    }} else {{
                        const scoreVal = a.score !== null && a.score !== undefined ? parseFloat(a.score) : 0;
                        if (!isNaN(scoreVal) && scoreVal >= 4.5 && diffVal !== null && diffVal >= 500) {{
                            finalResult = '〇';
                        }} else {{
                            finalResult = '×';
                        }}
                    }}
                }}
                
                const resultBadge = finalResult === '〇' ? '<span class="px-2 py-0.5 bg-emerald-500/20 text-emerald-400 text-[10px] rounded-full font-bold">〇</span>' : 
                                   (finalResult === '×' ? '<span class="px-2 py-0.5 bg-rose-500/20 text-rose-400 text-[10px] rounded-full font-bold">×</span>' : '-');
                const rankBadge = extractRankBadge(a.reason);
                return `
                    <td class="py-2.5 px-2 text-slate-200 font-semibold">${{a.number}}番 (${{a.name}})</td>
                    <td class="py-2.5 px-2 text-center">${{rankBadge}}</td>
                    <td class="py-2.5 px-2 font-mono ${{diffClass}} text-right">${{diffText}}</td>
                    <td class="py-2.5 px-2 text-slate-300 font-mono text-center">${{scoreText}}</td>
                    <td class="py-2.5 px-2 text-center">${{resultBadge}}</td>
                `;
            }};

            if (jugAnswers.length === 0) {{
                ansJugglerBody.innerHTML = '<tr><td colspan="5" class="py-3 text-center text-slate-500">答え合わせはありません。</td></tr>';
            }} else {{
                jugAnswers.forEach(a => {{
                    const tr = document.createElement('tr');
                    tr.className = 'hover:bg-slate-800/30';
                    tr.innerHTML = renderAnswerRow(a, true);
                    ansJugglerBody.appendChild(tr);
                }});
            }}

            if (otherAnswers.length === 0) {{
                ansOthersBody.innerHTML = '<tr><td colspan="5" class="py-3 text-center text-slate-500">答え合わせはありません。</td></tr>';
            }} else {{
                otherAnswers.forEach(a => {{
                    const tr = document.createElement('tr');
                    tr.className = 'hover:bg-slate-800/30';
                    tr.innerHTML = renderAnswerRow(a, false);
                    ansOthersBody.appendChild(tr);
                }});
            }}

            // 5. Populate Juggler High Setting List (Score >= 4.5) for targetDate
            const jugHighScoreBody = document.getElementById('juggler-high-score-body');
            const jugHighCountSpan = document.getElementById('jug-high-count');
            if (jugHighScoreBody) {{
                jugHighScoreBody.innerHTML = '';
                const jugHighRecords = filteredRecords.filter(r => isJugglerMachine(r.name) && (r.score >= 4.5)).sort((a, b) => b.score - a.score);
                if (jugHighCountSpan) jugHighCountSpan.textContent = `${{jugHighRecords.length}}台`;
                
                if (jugHighRecords.length === 0) {{
                    jugHighScoreBody.innerHTML = '<tr><td colspan="5" class="py-3 text-center text-slate-500">該当する高設定挙動台(4.5以上)はありません。</td></tr>';
                }} else {{
                    jugHighRecords.forEach(r => {{
                        const tr = document.createElement('tr');
                        tr.className = 'hover:bg-slate-800/30';
                        const scoreVal = r.score ? parseFloat(r.score).toFixed(1) : '-';
                        const diffText = HAS_DIFF_COINS ? (r.diff >= 0 ? '+' + r.diff : r.diff) + ' 枚' : (r.comb_ratio || '-');
                        const diffClass = HAS_DIFF_COINS ? (r.diff >= 0 ? 'text-emerald-400 font-bold' : 'text-rose-400') : 'text-slate-300';
                        
                        tr.innerHTML = `
                            <td class="py-2 px-1 font-semibold text-slate-200">${{r.number}}番 <span class="text-[10px] text-slate-400 font-normal">(${{r.name}})</span></td>
                            <td class="py-2 px-1 text-right text-slate-300">${{(r.games || 0).toLocaleString()}}G <span class="text-[10px] text-slate-400">(${{r.rb_ratio || '-'}})</span></td>
                            <td class="py-2 px-1 text-right ${{diffClass}}">${{diffText}}</td>
                            <td class="py-2 px-1 text-center font-bold text-amber-400">${{scoreVal}} 🌟</td>
                        `;
                        jugHighScoreBody.appendChild(tr);
                    }});
                }}
            }}

            // 6. Populate Store Top 5 Performance (Diff Coins or Games/Score)
            const topPerfBody = document.getElementById('top-performance-body');
            if (topPerfBody) {{
                topPerfBody.innerHTML = '';
                const topRecords = [...filteredRecords].sort((a, b) => HAS_DIFF_COINS ? (b.diff - a.diff) : (b.games - a.games)).slice(0, 5);
                
                if (topRecords.length === 0) {{
                    topPerfBody.innerHTML = '<tr><td colspan="3" class="py-3 text-center text-slate-500">実績データがありません。</td></tr>';
                }} else {{
                    topRecords.forEach((r, idx) => {{
                        const tr = document.createElement('tr');
                        tr.className = 'hover:bg-slate-800/30';
                        const rankBadge = idx === 0 ? '🥇' : (idx === 1 ? '🥈' : (idx === 2 ? '🥉' : `${{idx + 1}}.`));
                        const valText = HAS_DIFF_COINS ? ((r.diff >= 0 ? '+' : '') + r.diff.toLocaleString() + ' 枚') : (r.score ? r.score.toFixed(1) + ' 🌟' : '-');
                        const valClass = HAS_DIFF_COINS ? (r.diff >= 0 ? 'text-emerald-400 font-bold' : 'text-rose-400') : 'text-amber-400 font-bold';
                        
                        tr.innerHTML = `
                            <td class="py-2 px-1 font-semibold text-slate-200">${{rankBadge}} ${{r.number}}番 <span class="text-[10px] text-slate-400 font-normal">(${{r.name}})</span></td>
                            <td class="py-2 px-1 text-right text-slate-300">${{(r.games || 0).toLocaleString()}} G</td>
                            <td class="py-2 px-1 text-right ${{valClass}}">${{valText}}</td>
                        `;
                        topPerfBody.appendChild(tr);
                    }});
                }}
            }}

            // 5. Populate Raw Data Table
            populateRawTable(filteredRecords);

            // Setup Search
            const searchInput = document.getElementById('search-input');
            const newSearchInput = searchInput.cloneNode(true);
            searchInput.parentNode.replaceChild(newSearchInput, searchInput);
            newSearchInput.value = '';
            newSearchInput.addEventListener('input', (e) => {{
                const q = e.target.value.toLowerCase();
                const searched = filteredRecords.filter(r => 
                    r.name.toLowerCase().includes(q) || String(r.number).includes(q)
                );
                populateRawTable(searched);
            }});

            // 6. Draw Chart
            drawMachineChart(filteredRecords);
        }}

        function populateRawTable(data) {{
            const tbody = document.getElementById('raw-data-table-body');
            tbody.innerHTML = '';
            if (data.length === 0) {{
                tbody.innerHTML = '<tr><td colspan="9" class="py-8 text-center text-slate-500">データが見つかりません。</td></tr>';
                return;
            }}
            data.forEach(r => {{
                const tr = document.createElement('tr');
                tr.className = 'hover:bg-slate-800/30 transition-colors';
                const diffClass = r.diff >= 0 ? 'text-emerald-400 font-bold' : 'text-rose-400';
                const diffText = r.diff >= 0 ? '+' + r.diff : r.diff;
                tr.innerHTML = `
                    <td class="py-3 px-2 text-slate-400 font-semibold font-mono">${{r.number}}</td>
                    <td class="py-3 px-2 font-bold text-slate-200">${{r.name}}</td>
                    <td class="py-3 px-2 text-slate-300 font-mono">${{r.games}} G</td>
                    <td class="py-3 px-2 font-mono ${{diffClass}}">${{diffText}} 枚</td>
                    <td class="py-3 px-2 text-slate-400 font-mono">${{r.bb}}</td>
                    <td class="py-3 px-2 text-slate-400 font-mono">${{r.rb}}</td>
                    <td class="py-3 px-2 text-slate-300 font-mono">${{r.comb_ratio}}</td>
                    <td class="py-3 px-2 text-slate-400 font-mono">${{r.bb_ratio}}</td>
                    <td class="py-3 px-2 text-slate-400 font-mono">${{r.rb_ratio}}</td>
                `;
                tbody.appendChild(tr);
            }});
        }}

                                        function getRankFromReason(reason) {{
            if (!reason) return 'B';
            if (reason.includes('テスト')) return 'TEST';
            
            const match = reason.match(/(?:推奨)?([SAB])(?:ランク)?/i);
            if (match) {{
                return match[1].toUpperCase();
            }}
            
            if (reason.includes('大本命') || reason.includes('最優先') || reason.includes('超有力')) return 'S';
            if (reason.includes('本命') || reason.includes('有力') || reason.includes('対抗') || reason.includes('高期待')) return 'A';
            
            return 'B';
        }}

        function calculateOverallAIAccuracy(targetDate) {{
            const deduped = deduplicatePredictions(rawPredictions);
            const evaluated = deduped.filter(p => {{
                if (p.result !== '〇' && p.result !== '×') return false;
                if (!p.reason || p.reason.includes('テスト')) return false;
                if (targetDate && p.date > targetDate) return false;
                
                const rank = getRankFromReason(p.reason);
                return rank === 'S' || rank === 'A';
            }});
            
            const statElem = document.getElementById('stat-ai-accuracy');
            if (!statElem) return;
            
            if (evaluated.length === 0) {{
                statElem.textContent = '- %';
                return;
            }}
            const correct = evaluated.filter(p => p.result === '〇').length;
            const accuracy = Math.round((correct / evaluated.length) * 100);
            statElem.textContent = accuracy + '% (' + correct + '/' + evaluated.length + ')';
        }}

        function drawMachineChart(records) {{
            const machineData = {{}};
            records.forEach(r => {{
                if (!machineData[r.name]) {{
                    machineData[r.name] = {{ totalDiff: 0, count: 0 }};
                }}
                machineData[r.name].totalDiff += r.diff;
                machineData[r.name].count += 1;
            }});

            const chartItems = Object.keys(machineData).map(k => ({{
                name: k,
                avgDiff: Math.round(machineData[k].totalDiff / machineData[k].count)
            }})).sort((a, b) => b.avgDiff - a.avgDiff).slice(0, 10);

            const labels = chartItems.map(item => item.name);
            const dataVals = chartItems.map(item => item.avgDiff);

            const ctx = document.getElementById('chart-machines').getContext('2d');
            if (machineChart) {{
                machineChart.destroy();
            }}

            machineChart = new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: labels,
                    datasets: [{{
                        label: '平均差枚数 (枚)',
                        data: dataVals,
                        backgroundColor: dataVals.map(v => v >= 0 ? 'rgba(52, 211, 153, 0.6)' : 'rgba(248, 113, 113, 0.6)'),
                        borderColor: dataVals.map(v => v >= 0 ? 'rgba(52, 211, 153, 1)' : 'rgba(248, 113, 113, 1)'),
                        borderWidth: 1.5,
                        borderRadius: 6
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    layout: {{
                        padding: {{
                            top: 25
                        }}
                    }},
                    plugins: {{
                        legend: {{ display: false }},
                        datalabels: {{
                            anchor: 'end',
                            align: 'top',
                            color: '#e2e8f0',
                            font: {{
                                weight: 'bold',
                                size: 10
                            }},
                            formatter: function(value) {{
                                return (value >= 0 ? '+' : '') + value;
                            }}
                        }}
                    }},
                    scales: {{
                        y: {{
                            grid: {{ color: 'rgba(255, 255, 255, 0.05)' }},
                            ticks: {{ color: '#94a3b8' }}
                        }},
                        x: {{
                            grid: {{ display: false }},
                            ticks: {{ 
                                color: '#94a3b8',
                                font: {{ size: 10 }}
                            }}
                        }}
                    }}
                }}
            }});
        }}
    </script>
</body>
</html>
"""
    
    output_html_path = os.path.join(os.path.dirname(excel_path), f"分析ダッシュボード_{store_name}.html")
    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"HTML Interactive Dashboard generated successfully: {output_html_path}")
    
    # Also save as index.html for root page on GitHub Pages
    index_html_path = os.path.join(os.path.dirname(excel_path), "index.html")
    with open(index_html_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"HTML index.html generated successfully for GitHub Pages.")

def deploy_to_github():
    """
    Deploys the generated HTML dashboard and raw summaries to GitHub Pages automatically
    if the local directory is initialized as a Git repository.
    """
    print("\nDeploying dashboard to GitHub Pages...")
    try:
        if not os.path.exists(".git"):
            print("Notice: Git repository not initialized. Skipping auto-deploy.")
            return
            
        # Add generated html and rich summaries
        os.system("git add 分析ダッシュボード_*.html index.html data_input/ai_summaries_rich.json")
        
        # Commit changes
        commit_msg = f"Auto-update dashboard: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
        os.system(f'git commit -m "{commit_msg}"')
        
        # Push to remote (Try main first, then master)
        res = os.system("git push origin main")
        if res != 0:
            os.system("git push origin master")
            
        print("Successfully deployed to GitHub Pages! It will be live in a few seconds.")
    except Exception as e:
        print(f"Warning: Failed to deploy to GitHub: {e}")

def main():
    print("=== Pachislot Data Analyzer Auto Tool ===")
    config = load_config()
    
    api_key = config.get("GEMINI_API_KEY")
    if not api_key or api_key == "YOUR_GEMINI_API_KEY_HERE":
        print("Error: Gemini API key is missing in config.json. Please update the key.")
        sys.exit(1)
        
    allowed_exts = [".txt", ".html", ".htm"]
    files = [f for f in os.listdir("data_input") if os.path.splitext(f)[1].lower() in allowed_exts]
    if not files:
        print("\nNo input data files found in 'data_input/' folder.")
        print("Please place a text or downloaded HTML file (e.g. 'A店_20260714.html') inside 'data_input/' and try again.")
        sys.exit(0)
        
    print("\nSelect the file to process:")
    for idx, f in enumerate(files):
        print(f"{idx + 1}: {f}")
        
    if len(files) == 1:
        print(f"\nOnly one file found: '{files[0]}'. Automatically selecting it.")
        sel = 0
    else:
        # Check if stdin is interactive
        if sys.stdin.isatty():
            try:
                sel = int(input("\nEnter selection number: ")) - 1
                if sel < 0 or sel >= len(files):
                    raise ValueError()
            except ValueError:
                print("Invalid selection.")
                sys.exit(1)
        else:
            print("\nNon-interactive shell detected. Defaulting to 1st file.")
            sel = 0
        
    selected_file = files[sel]
    filepath = os.path.join("data_input", selected_file)
    
    filename_wo_ext = os.path.splitext(selected_file)[0]
    
    # 1. Automatic Store Name Detection
    store_name = None
    registered_stores = list(config.get("STORES", {}).keys())
    
    # Check if filename contains any registered store key
    for s_key in registered_stores:
        if s_key.lower() in filename_wo_ext.lower():
            store_name = s_key
            break
            
    # Fallback to single registered store if only 1 exists in config
    if not store_name and len(registered_stores) == 1:
        store_name = registered_stores[0]

    # 2. Automatic Date Detection (Search for 8-digit date YYYYMMDD anywhere in filename)
    date_match = re.search(r'(\d{8})', filename_wo_ext)
    if date_match:
        target_date_raw = date_match.group(1)
    else:
        target_date_raw = None

    # Fallback prompt only if detection fails completely
    if not store_name:
        store_name = input("Enter Store Name (e.g. 123 / トワーズ大和深見): ")
    if not target_date_raw:
        target_date_raw = input("Enter Date (e.g. 20260401): ")
        
    try:
        target_date = datetime.datetime.strptime(target_date_raw, "%Y%m%d").strftime("%Y/%m/%d")
    except ValueError:
        print(f"Error: Invalid date format '{target_date_raw}'. Expected YYYYMMDD.")
        sys.exit(1)
        
    store_info = config.get("STORES", {}).get(store_name)
    if not store_info:
        print(f"Error: Store '{store_name}' is not registered in config.json.")
        sys.exit(1)
        
    excel_file = store_info.get("excel_file")
    print(f"\nTarget Store: {store_name}")
    print(f"Target Date: {target_date}")
    print(f"Target Excel: {excel_file}")
    
    ext = os.path.splitext(selected_file)[1].lower()
    if ext in [".html", ".htm"]:
        parsed_data = parse_html_data(filepath)
    else:
        parsed_data = parse_text_data(filepath)
        
    if not parsed_data:
        print("Error: No data rows parsed. Check the input file format.")
        sys.exit(1)
        
    update_excel_data(excel_file, target_date, parsed_data)
    
    print("\nExtracting Excel metrics for AI context...")
    context = prepare_ai_context(excel_file, target_date)
    
    print("\nRunning Gemini AI analysis...")
    ai_text = run_gemini_analysis(api_key, context, target_date, excel_file)
    
    ref_out = os.path.join("data_input", f"{filename_wo_ext}_ai_report.md")
    with open(ref_out, "w", encoding="utf-8") as f:
        f.write(ai_text)
    print(f"AI raw report saved to: {ref_out}")
    
    print("\nWriting AI recommendations back to Excel...")
    write_ai_results_to_excel(excel_file, target_date, ai_text)
    
    # 4. Generate the fully interactive HTML Dashboard for instant check
    generate_html_dashboard(excel_file, store_name, has_diff_coins=store_info.get('has_diff_coins', True))
    
    # 5. Automatically deploy to GitHub Pages
    deploy_to_github()
    
    os.makedirs(os.path.join("data_input", "processed"), exist_ok=True)
    os.rename(filepath, os.path.join("data_input", "processed", selected_file))
    print(f"\n=== Process complete! Input file moved to processed/ backup folder. ===")

if __name__ == "__main__":
    main()